from io import BytesIO

import pandas as pd
from flask import Flask, jsonify, render_template, request, send_file, send_from_directory

from parsers.csv_parser import parse_csv
from parsers.excel_parser import parse_excel
from services.codechef_service import get_cc_summary, get_latest_cc_contests
from services.codeforces_service import get_cf_summary, get_latest_cf_contests
from services.contest_scheduler import contest_scheduler
from services.contest_service import contest_service
from services.leetcode_service import find_latest_lc_contest, get_lc_summary, get_latest_lc_contests
from services.notification_service import notification_manager
from services.student_service import StudentService
from services.topper_service import compute_topper
from utils.date_utils import get_export_filename
from utils.excel_utils import create_excel_file

try:
    contest_scheduler.start()
except Exception as _e:
    pass


app = Flask(__name__)


cache_tables = {
    "codeforces": [],
    "codechef": [],
    "leetcode": [],
}


def _clean_text(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def _selected_platforms(form):
    selected = []
    for platform in ("codeforces", "codechef", "leetcode"):
        if form.get(f"platform_{platform}"):
            selected.append(platform)
    return selected


def _row_value(row, *keys):
    for key in keys:
        value = _clean_text(row.get(key))
        if value:
            return value
    return ""


def _normalize_rows(rows):
    normalized = []
    for row in rows:
        normalized.append(
            {
                "name": _row_value(row, "name", "studentName"),
                "studentName": _row_value(row, "studentName", "name"),
                "register_no": _row_value(row, "register_no", "registerNo"),
                "registerNo": _row_value(row, "registerNo", "register_no"),
                "department": _row_value(row, "department", "dept"),
                "codeforces": _row_value(row, "codeforces"),
                "codechef": _row_value(row, "codechef"),
                "leetcode": _row_value(row, "leetcode"),
            }
        )
    return normalized


def _rows_from_form(form):
    row = {
        "name": _clean_text(form.get("name")),
        "studentName": _clean_text(form.get("name")),
        "register_no": _clean_text(form.get("register_no")),
        "registerNo": _clean_text(form.get("register_no")),
        "department": _clean_text(form.get("department")),
        "codeforces": _clean_text(form.get("codeforces")),
        "codechef": _clean_text(form.get("codechef")),
        "leetcode": _clean_text(form.get("leetcode")),
    }
    if not row["name"] and not row["register_no"]:
        return []
    return [row]


def _load_rows(uploaded_file):
    filename = (uploaded_file.filename or "").lower()
    if filename.endswith(".csv"):
        return parse_csv(uploaded_file)
    if filename.endswith((".xlsx", ".xls")):
        return parse_excel(uploaded_file)
    return None


def _analyze_rows(rows, selected_platforms, lc_targets=None, cc_targets=None, cf_targets=None):
    """
    Returns dict mapping platform names to a list of contest table blocks:
    {
        "leetcode": [
            {"contest": "Weekly Contest 400", "date": "2026-07-26", "rows": [...]},
            ...
        ],
        ...
    }
    """
    tables = {"codeforces": [], "codechef": [], "leetcode": []}

    if "leetcode" in selected_platforms and not lc_targets:
        latest_title, latest_time = find_latest_lc_contest(rows)
        if latest_title:
            lc_targets = [{"title": latest_title, "startTime": latest_time}]
        else:
            lc_targets = [{"title": None, "startTime": None}]

    if "codechef" in selected_platforms and not cc_targets:
        cc_targets = [{"title": None, "date": None}]

    if "codeforces" in selected_platforms and not cf_targets:
        cf_targets = [{"title": None, "id": None, "date": None}]

    if "codeforces" in selected_platforms:
        for cf_t in (cf_targets or [{"title": None, "id": None, "date": None}]):
            c_title = cf_t.get("title") or (str(cf_t.get("id")) if cf_t.get("id") else "General Summary")
            c_rows = []
            c_idx = 1
            for row in rows:
                name = _clean_text(row.get("name") or row.get("studentName"))
                regno = _clean_text(row.get("register_no") or row.get("registerNo"))
                dept = _clean_text(row.get("department"))
                if not name:
                    continue
                handle = _clean_text(row.get("codeforces"))
                if handle:
                    c_rows.append(
                        get_cf_summary(
                            c_idx, name, regno, dept, handle,
                            target_contest_id=cf_t.get("id"),
                            target_contest_date=cf_t.get("date"),
                            target_contest_title=cf_t.get("title")
                        )
                    )
                    c_idx += 1
            tables["codeforces"].append({"contest": c_title, "date": cf_t.get("date"), "rows": c_rows})

    if "codechef" in selected_platforms:
        for cc_t in (cc_targets or [{"title": None, "date": None}]):
            c_title = cc_t.get("title") or "General Summary"
            c_rows = []
            c_idx = 1
            for row in rows:
                name = _clean_text(row.get("name") or row.get("studentName"))
                regno = _clean_text(row.get("register_no") or row.get("registerNo"))
                dept = _clean_text(row.get("department"))
                if not name:
                    continue
                handle = _clean_text(row.get("codechef"))
                if handle:
                    c_rows.append(
                        get_cc_summary(
                            c_idx, name, regno, dept, handle,
                            target_contest_title=cc_t.get("title"),
                            target_contest_date=cc_t.get("date")
                        )
                    )
                    c_idx += 1
            tables["codechef"].append({"contest": c_title, "date": cc_t.get("date"), "rows": c_rows})

    if "leetcode" in selected_platforms:
        for lc_t in (lc_targets or [{"title": None, "startTime": None}]):
            c_title = lc_t.get("title") or "General Summary"
            c_rows = []
            c_idx = 1
            for row in rows:
                name = _clean_text(row.get("name") or row.get("studentName"))
                regno = _clean_text(row.get("register_no") or row.get("registerNo"))
                dept = _clean_text(row.get("department"))
                if not name:
                    continue
                handle = _clean_text(row.get("leetcode"))
                if handle:
                    c_rows.append(
                        get_lc_summary(
                            c_idx, name, regno, dept, handle,
                            lc_t.get("title"),
                            lc_t.get("startTime")
                        )
                    )
                    c_idx += 1
            tables["leetcode"].append({"contest": c_title, "rows": c_rows})

    return tables


def _clean_row_dict(row):
    """Clean row dictionary to replace None/NaN with 'AB' or empty string and ensure numeric values don't turn into floats."""
    cleaned = {}
    for k, v in row.items():
        if v is None or v == "" or pd.isna(v):
            cleaned[k] = "AB"
        elif isinstance(v, float) and v.is_integer():
            cleaned[k] = int(v)
        else:
            cleaned[k] = v
    return cleaned


def _combined_export_frame(tables, requested_platform=None):
    """Build a combined pandas DataFrame for CSV export."""
    frames = []

    target_keys = [requested_platform] if (requested_platform and requested_platform in tables) else ["codeforces", "codechef", "leetcode"]

    for platform in target_keys:
        contest_blocks = tables.get(platform, [])
        for block in contest_blocks:
            rows = block.get("rows", [])
            if not rows:
                continue
            cleaned_rows = [_clean_row_dict(r) for r in rows]
            frame = pd.DataFrame(cleaned_rows)
            if len(target_keys) > 1:
                frame.insert(0, "Platform", platform.capitalize())
            frames.append(frame)

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined = combined.fillna("AB")
    return combined


def _auto_fit_columns(worksheet):
    """Auto-adjust worksheet column widths and apply header styling."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11, color="0F172A")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in worksheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)


def _tables_to_excel_stream(tables, requested_platform=None):
    """Build an Excel file stream with dedicated contest worksheets and styling."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        workbook = writer.book

        target_keys = [requested_platform] if (requested_platform and requested_platform in tables) else ["codeforces", "codechef", "leetcode"]

        # Combined sheet first
        combined_frame = _combined_export_frame(tables, requested_platform=requested_platform)
        if not combined_frame.empty:
            combined_frame.to_excel(writer, sheet_name="Combined Results", index=False)
            ws_comb = writer.sheets["Combined Results"]
            _auto_fit_columns(ws_comb)

        # Dedicated sheet per contest table
        sheet_count = {}
        for platform in target_keys:
            contest_blocks = tables.get(platform, [])
            for block in contest_blocks:
                contest_title = block.get("contest") or platform.capitalize()
                rows = block.get("rows", [])
                if not rows:
                    continue

                base_name = f"{platform[:2].upper()} - {contest_title}"[:28]
                sheet_count[base_name] = sheet_count.get(base_name, 0) + 1
                sheet_name = base_name if sheet_count[base_name] == 1 else f"{base_name[:25]} ({sheet_count[base_name]})"

                cleaned_rows = [_clean_row_dict(r) for r in rows]
                df = pd.DataFrame(cleaned_rows)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                _auto_fit_columns(ws)

        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook["Sheet"]

    output.seek(0)
    return output


@app.route("/favicon.ico")
def favicon():
    return send_from_directory(app.static_folder, "favicon.ico", mimetype="image/vnd.microsoft.icon")


@app.route("/api/leetcode/contests", methods=["GET"])
def api_leetcode_contests():
    try:
        contests = get_latest_lc_contests(6)
        return jsonify({"contests": contests})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch LeetCode contests from API: {str(e)}", "contests": []}), 500


@app.route("/api/codechef/contests", methods=["GET"])
def api_codechef_contests():
    try:
        contests = get_latest_cc_contests(6)
        return jsonify({"contests": contests})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch CodeChef contests from API: {str(e)}", "contests": []}), 500


@app.route("/api/codeforces/contests", methods=["GET"])
def api_codeforces_contests():
    try:
        contests = get_latest_cf_contests(6)
        return jsonify({"contests": contests})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch Codeforces contests from API: {str(e)}", "contests": []}), 500


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "GET":
        return render_template("index.html")

    selected_platforms = _selected_platforms(request.form)
    if not selected_platforms:
        return render_template("index.html", error="Select at least one platform to track."), 400

    lc_targets = []
    if "leetcode" in selected_platforms:
        lc_vals = [c.strip() for c in request.form.getlist("leetcode_contest") if c and c.strip()]
        if not lc_vals and request.form.get("leetcode_contest"):
            lc_vals = [request.form.get("leetcode_contest").strip()]

        if not lc_vals:
            return render_template("index.html", error="Please select at least one LeetCode contest before fetching rankings."), 400

        try:
            contests = get_latest_lc_contests(15)
            for val in lc_vals:
                matched = next((c for c in contests if c["title"] == val or c["titleSlug"] == val), None)
                if matched:
                    lc_targets.append({"title": matched["title"], "startTime": matched["startTime"]})
                else:
                    lc_targets.append({"title": val, "startTime": 0})
        except Exception as e:
            return render_template("index.html", error=f"Failed to fetch LeetCode contests from API: {str(e)}"), 500

    cc_targets = []
    if "codechef" in selected_platforms:
        cc_vals = [c.strip() for c in request.form.getlist("codechef_contest") if c and c.strip()]
        if not cc_vals and request.form.get("codechef_contest"):
            cc_vals = [request.form.get("codechef_contest").strip()]

        if cc_vals:
            try:
                cc_contests = get_latest_cc_contests(15)
                for val in cc_vals:
                    matched_cc = next((c for c in cc_contests if c["title"] == val or c["code"] == val), None)
                    if matched_cc:
                        cc_targets.append({"title": matched_cc["title"], "date": matched_cc["date"]})
                    else:
                        cc_targets.append({"title": val, "date": None})
            except Exception:
                for val in cc_vals:
                    cc_targets.append({"title": val, "date": None})

    cf_targets = []
    if "codeforces" in selected_platforms:
        cf_vals = [c.strip() for c in request.form.getlist("codeforces_contest") if c and c.strip()]
        if not cf_vals and request.form.get("codeforces_contest"):
            cf_vals = [request.form.get("codeforces_contest").strip()]

        if cf_vals:
            try:
                cf_contests = get_latest_cf_contests(15)
                for val in cf_vals:
                    matched_cf = next((c for c in cf_contests if c["title"] == val or str(c.get("id")) == val), None)
                    if matched_cf:
                        cf_targets.append({"title": matched_cf["title"], "id": matched_cf.get("id"), "date": matched_cf.get("date")})
                    else:
                        cf_targets.append({"title": val, "id": val, "date": None})
            except Exception:
                for val in cf_vals:
                    cf_targets.append({"title": val, "id": val, "date": None})

    uploaded_file = request.files.get("csvfile")
    rows = []

    if uploaded_file and uploaded_file.filename:
        loaded_rows = _load_rows(uploaded_file)
        if loaded_rows is None:
            return render_template("index.html", error="Upload a CSV or Excel file."), 400
        rows = loaded_rows
    else:
        rows = _rows_from_form(request.form)

    if not rows:
        return render_template("index.html", error="Add a student or upload a file with rows to analyze."), 400

    rows = _normalize_rows(rows)

    if not any(_clean_text(row.get(platform)) for row in rows for platform in selected_platforms):
        return render_template(
            "index.html",
            error="Provide at least one platform ID for the selected platforms.",
        ), 400

    tables = _analyze_rows(rows, selected_platforms, lc_targets, cc_targets, cf_targets)

    global cache_tables
    cache_tables = tables

    student_count = len(rows)
    platforms_str = ", ".join([p.capitalize() for p in selected_platforms])
    if uploaded_file and uploaded_file.filename:
        toast_title = "File Upload & Evaluation Completed 🚀"
        toast_msg = f"Successfully uploaded '{uploaded_file.filename}' and evaluated stats for {student_count} student records across {platforms_str}."
    else:
        toast_title = "Student Stats Evaluation Completed 🎯"
        toast_msg = f"Successfully evaluated stats for {student_count} student records across {platforms_str}."

    notification_manager.send_notification(
        user_id="default_user",
        title=toast_title,
        message=toast_msg,
        n_type="fetch_complete"
    )

    return render_template(
        "results.html",
        codeforces=tables["codeforces"],
        codechef=tables["codechef"],
        leetcode=tables["leetcode"],
        selected_platforms=selected_platforms,
        completion_toast={"title": toast_title, "message": toast_msg, "icon": "🚀" if uploaded_file else "🎯"}
    )


@app.route("/download")
def download():
    export_format = request.args.get("format", "xlsx").lower()
    requested_platform = request.args.get("platform", "").lower().strip()

    has_data = any(
        any(b.get("rows") for b in cache_tables.get(key, []))
        for key in ("codeforces", "codechef", "leetcode")
    )
    if not has_data:
        return "No data to download.", 404

    active_platforms = [k for k, v in cache_tables.items() if v]

    if requested_platform and cache_tables.get(requested_platform):
        platform_name = requested_platform
        export_tables = {requested_platform: cache_tables[requested_platform]}
    else:
        if len(active_platforms) == 1:
            platform_name = active_platforms[0]
        else:
            platform_name = requested_platform if requested_platform else "platstat"
        export_tables = cache_tables

    filename = get_export_filename(platform_name=platform_name, extension=export_format)

    if export_format == "csv":
        frame = _combined_export_frame(export_tables, requested_platform=requested_platform if requested_platform in cache_tables else None)
        output = BytesIO()
        output.write(frame.to_csv(index=False).encode("utf-8-sig"))
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="text/csv; charset=utf-8",
        )

    excel_file = _tables_to_excel_stream(export_tables, requested_platform=requested_platform if requested_platform in cache_tables else None)
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/classes/<class_id>/students/export", methods=["GET"])
def export_class_students(class_id):
    export_format = request.args.get("format", "xlsx").lower()
    platform_name = request.args.get("platform", "platstat").lower().strip()
    student_service = StudentService()
    excel_stream = student_service.export_students_to_excel(class_id)
    if not excel_stream:
        return "No student records found.", 404
    filename = get_export_filename(platform_name=platform_name, extension=export_format)
    if export_format == "csv":
        df = pd.read_excel(excel_stream)
        csv_output = BytesIO()
        csv_output.write(df.to_csv(index=False).encode("utf-8"))
        csv_output.seek(0)
        return send_file(
            csv_output,
            as_attachment=True,
            download_name=filename,
            mimetype="text/csv",
        )
    return send_file(
        excel_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/topper", methods=["GET", "POST"])
def topper():
    if request.method == "GET":
        return render_template("topper.html", result=None, error=None, view_mode="both")

    error = None
    result = None
    view_mode = (request.form.get("view_mode") or "both").lower()

    try:
        file = request.files.get("sheet")
        platform = _clean_text(request.form.get("platform"))
        month_raw = _clean_text(request.form.get("month"))

        if not file or not file.filename:
            error = "Upload a CSV or Excel file."
            return render_template("topper.html", result=None, error=error, view_mode=view_mode)

        if not platform:
            error = "Select a platform."
            return render_template("topper.html", result=None, error=error, view_mode=view_mode)

        if not month_raw:
            error = "Select a month."
            return render_template("topper.html", result=None, error=error, view_mode=view_mode)

        month = int(month_raw)
        filename = file.filename.lower()
        if filename.endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)

        ranked = compute_topper(df, platform, month)
        if ranked is None or ranked.empty:
            error = "No records found for the selected month."
        else:
            result = {}
            if view_mode in ("5", "top5", "both"):
                result["top5"] = ranked.head(5).to_dict("records")
            if view_mode in ("10", "top10", "both"):
                result["top10"] = ranked.head(10).to_dict("records")

            month_names = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}
            month_label = month_names.get(month, f"Month {month}")
            toast_title = "Topper Calculation Completed 🏆"
            toast_msg = f"Calculated top performers for {platform.capitalize()} ({month_label})."
            notification_manager.send_notification(
                user_id="default_user",
                title=toast_title,
                message=toast_msg,
                n_type="topper_complete"
            )
            return render_template(
                "topper.html",
                result=result,
                error=error,
                view_mode=view_mode,
                completion_toast={"title": toast_title, "message": toast_msg, "icon": "🏆"}
            )

    except Exception as exc:
        error = str(exc)

    return render_template("topper.html", result=result, error=error, view_mode=view_mode)


# ----------------------------------------------------
# Contest Center & Reminder Module Endpoints
# ----------------------------------------------------
@app.route("/contests")
def contest_center():
    return render_template("contests.html")


@app.route("/notifications")
def notification_center():
    return render_template("notifications.html")


@app.route("/api/contests", methods=["GET"])
def get_contests():
    platform = request.args.get("platform")
    search = request.args.get("search")
    favorites_only = request.args.get("favorites") == "true"
    sort_by = request.args.get("sort", "nearest")
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 50))
    user_id = request.args.get("user_id", "default_user")

    items, total = contest_service.get_upcoming_contests(
        platform=platform,
        search=search,
        favorites_only=favorites_only,
        sort_by=sort_by,
        user_id=user_id,
        page=page,
        page_size=page_size
    )
    return jsonify({"contests": items, "total": total, "page": page, "pageSize": page_size})


@app.route("/api/contests/sync", methods=["POST"])
def sync_contests():
    res = contest_service.sync_contests()
    synced_cnt = res.get("syncedCount", 0) if isinstance(res, dict) else 0
    toast_title = "Contest Sync Completed 🔄"
    toast_msg = f"Synced latest competitive programming contests ({synced_cnt} active/upcoming)."
    notification_manager.send_notification(
        user_id="default_user",
        title=toast_title,
        message=toast_msg,
        n_type="sync_complete"
    )
    if isinstance(res, dict):
        res["toast"] = {"title": toast_title, "message": toast_msg, "icon": "🔄"}
    return jsonify(res)


@app.route("/api/contests/<contest_id>/favorite", methods=["POST"])
def toggle_contest_favorite(contest_id):
    data = request.get_json(silent=True) or {}
    platform = data.get("platform", "")
    favorite = data.get("favorite", True)
    user_id = data.get("user_id", "default_user")
    contest_service.toggle_favorite(user_id, contest_id, platform, favorite=favorite)
    return jsonify({"status": "success", "contestId": contest_id, "favorite": favorite})


@app.route("/api/contests/<contest_id>/subscribe", methods=["POST"])
def subscribe_contest_reminder(contest_id):
    data = request.get_json(silent=True) or {}
    platform = data.get("platform", "")
    intervals = data.get("intervals", ["1h"])
    user_id = data.get("user_id", "default_user")
    contest_service.subscribe_reminder(user_id, contest_id, platform, intervals=intervals)
    return jsonify({"status": "success", "contestId": contest_id, "intervals": intervals})


@app.route("/api/contests/<contest_id>/unsubscribe", methods=["POST"])
def unsubscribe_contest_reminder(contest_id):
    data = request.get_json(silent=True) or {}
    user_id = data.get("user_id", "default_user")
    contest_service.unsubscribe_reminder(user_id, contest_id)
    return jsonify({"status": "success", "contestId": contest_id})


@app.route("/api/user/reminders", methods=["GET"])
def get_user_reminders():
    user_id = request.args.get("user_id", "default_user")
    reminders = contest_service.reminder_repo.get_user_reminders(user_id)
    return jsonify({"reminders": reminders})


@app.route("/api/notifications", methods=["GET"])
def get_notifications():
    user_id = request.args.get("user_id", "default_user")
    items, unread = notification_manager.get_user_notifications(user_id)
    return jsonify({"notifications": items, "unreadCount": unread})


@app.route("/api/notifications/<notification_id>/read", methods=["POST"])
def mark_notification_read(notification_id):
    user_id = request.args.get("user_id", "default_user")
    notification_manager.mark_read(user_id, notification_id)
    return jsonify({"status": "success"})


@app.route("/api/notifications/clear", methods=["POST"])
def clear_notifications():
    user_id = request.args.get("user_id", "default_user")
    notification_manager.clear_all(user_id)
    return jsonify({"status": "success"})


@app.route("/api/dashboard/contest-widget", methods=["GET"])
def get_dashboard_contest_widget():
    user_id = request.args.get("user_id", "default_user")
    summary = contest_service.get_dashboard_contest_summary(user_id)
    return jsonify(summary)


@app.errorhandler(500)
def internal_error(_):
    return render_template("index.html", error="A critical error occurred. Please check your input or try again later."), 500


@app.errorhandler(Exception)
def handle_exception(e):
    if hasattr(e, "code") and e.code < 500:
        return e
    return render_template("index.html", error=f"Unexpected error: {str(e)}"), 500


if __name__ == "__main__":
    app.run(debug=True)
